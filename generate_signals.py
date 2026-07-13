import yfinance as yf
import pandas as pd
from datetime import datetime, timezone, timedelta
from openpyxl import load_workbook

IST = timezone(timedelta(hours=5, minutes=30))
generated_at_ist = datetime.now(timezone.utc).astimezone(IST).strftime('%Y-%m-%d %H:%M:%S IST')

ROLLING_WINDOW_MONTHS = 1.5
ROLLING_WINDOW_DAYS = int(21 * ROLLING_WINDOW_MONTHS)
Z_ENTRY_LEVELS = [2.0, 2.5, 3.0, 3.5]
MIN_AVG_PNL_PCT = 0.0  # only show if avg P&L at the applicable level is positive
OUTPUT_FILE = 'daily_signal_sheet.xlsx'

backtest_summary = pd.read_csv('backtest_reference.csv')

def parse_pair(pair_str):
    sector, tickers = pair_str.split(' | ')
    a, b = tickers.split('/')
    return sector, a, b

backtest_summary[['sector', 'stock_a', 'stock_b']] = backtest_summary['pair'].apply(
    lambda p: pd.Series(parse_pair(p)))

def get_applicable_z_level(current_abs_z, levels):
    applicable = [lvl for lvl in levels if lvl <= current_abs_z]
    return max(applicable) if applicable else None

all_tickers = pd.unique(backtest_summary[['stock_a', 'stock_b']].values.ravel())
prices = {}
for t in all_tickers:
    data = yf.download(t, period='4mo', auto_adjust=False, progress=False)['Close'].squeeze()
    if len(data) > 0:
        prices[t] = data

signal_rows = []
for _, row in backtest_summary.iterrows():
    a, b = row['stock_a'], row['stock_b']
    if a not in prices or b not in prices:
        continue
    pair_data = pd.DataFrame({a: prices[a], b: prices[b]}).dropna()
    if len(pair_data) < ROLLING_WINDOW_DAYS + 5:
        continue
    ratio = pair_data[a] / pair_data[b]
    roll_mean = ratio.rolling(ROLLING_WINDOW_DAYS).mean()
    roll_std = ratio.rolling(ROLLING_WINDOW_DAYS).std()
    z = (ratio - roll_mean) / roll_std
    latest_z = z.iloc[-1]

    applicable_level = get_applicable_z_level(abs(latest_z), Z_ENTRY_LEVELS)
    if applicable_level is None:
        continue  # below the smallest backtested threshold — no signal

    winrate_col = f'winrate_z{applicable_level}_%'
    n_col = f'n_z{applicable_level}'
    avgpnl_col = f'avgpnl_z{applicable_level}_%'

    if avgpnl_col not in row or pd.isna(row[avgpnl_col]):
        continue  # no backtest data at this level for this pair — can't evaluate, skip
    if row[avgpnl_col] < MIN_AVG_PNL_PCT:
        continue  # avg P&L at this level isn't positive — skip

    signal_rows.append({
        'pair': row['pair'],
        'signal_date': z.index[-1].date(),
        'current_ratio': round(ratio.iloc[-1], 4),
        'current_z': round(latest_z, 2),
        'applicable_z_level': applicable_level,
        'avgpnl_at_level_%': row[avgpnl_col],
        'winrate_at_level_%': row[winrate_col] if winrate_col in row and pd.notna(row[winrate_col]) else None,
        'n_signals_at_level': int(row[n_col]) if n_col in row and pd.notna(row[n_col]) else None,
        'signal_for_next_open': 'SHORT_A_LONG_B' if latest_z > 0 else 'LONG_A_SHORT_B'
    })

daily_signals = pd.DataFrame(signal_rows)
if len(daily_signals):
    daily_signals = daily_signals.sort_values('current_z', key=abs, ascending=False)

with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
    header_df = pd.DataFrame({
        'Info': [
            f'Sheet generated at: {generated_at_ist}',
            f'Filter: avg P&L >= {MIN_AVG_PNL_PCT}% at the applicable z-level (no minimum sample size — check n_signals_at_level yourself)',
            f'Active signals: {len(daily_signals)}'
        ]
    })
    header_df.to_excel(writer, sheet_name='SIGNALS_TODAY', index=False, header=False, startrow=0)
    daily_signals.to_excel(writer, sheet_name='SIGNALS_TODAY', index=False, startrow=4)
    backtest_summary.drop(columns=['stock_a', 'stock_b']).to_excel(
        writer, sheet_name='BACKTEST_REFERENCE', index=False)

wb = load_workbook(OUTPUT_FILE)
ws_signals = wb['SIGNALS_TODAY']
ws_ref = wb['BACKTEST_REFERENCE']

ref_pair_col = None
for col_cells in ws_ref.iter_cols(1, ws_ref.max_column):
    if col_cells[0].value == 'pair':
        ref_pair_col = col_cells[0].column_letter
        break

pair_to_row = {}
for r in range(2, ws_ref.max_row + 1):
    val = ws_ref[f'{ref_pair_col}{r}'].value
    if val:
        pair_to_row[val] = r

signal_header_row = 5
for r in range(signal_header_row + 1, ws_signals.max_row + 1):
    pair_cell = ws_signals[f'A{r}']
    pair_val = pair_cell.value
    if pair_val in pair_to_row:
        target_row = pair_to_row[pair_val]
        pair_cell.hyperlink = f"#'BACKTEST_REFERENCE'!A{target_row}"
        pair_cell.style = 'Hyperlink'

wb.save(OUTPUT_FILE)

print(f"Active signals today: {len(daily_signals)}")
